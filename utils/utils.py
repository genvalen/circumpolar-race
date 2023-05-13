import unicodedata
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side


def slugify(value, allow_unicode=False):
    """
    Taken from https://github.com/django/django/blob/master/django/utils/text.py
    Convert to ASCII if 'allow_unicode' is False. Convert spaces or repeated
    dashes to single dashes. Remove characters that aren't alphanumerics,
    underscores, or hyphens. Convert to lowercase. Also strip leading and
    trailing whitespace, dashes, and underscores.
    """
    value = str(value)
    if allow_unicode:
        value = unicodedata.normalize("NFKC", value)
    else:
        value = (
            unicodedata.normalize("NFKD", value)
            .encode("ascii", "ignore")
            .decode("ascii")
        )
    value = re.sub(r"[^\w\s-]", "", value.lower())
    return re.sub(r"[-\s]+", "-", value).strip("-_")


def style_spreadsheet(path, sheet_name):
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    end_of_table_row = sheet.max_row
    end_of_table_row_str = str(sheet.max_row)

    # Create style objects:
    # ---- Fonts
    font_bold = Font(bold=True)
    font_not_bold = Font(bold=False)

    # --- Borders
    thin_border = Side(border_style="thin")
    medium_border = Side(border_style="medium")
    no_border = Side(border_style=None)
    full_border_thin = Border(
        left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
    )
    full_border_medium = Border(
        left=medium_border, right=medium_border, top=medium_border, bottom=medium_border
    )
    top_border_thin = Border(
        left=no_border, right=no_border, top=thin_border, bottom=no_border
    )
    column_a_custom_border = Border(bottom=no_border)

    # Apply styles to cells
    for col in sheet.iter_cols(min_col=3, max_col=14, min_row=end_of_table_row):
        for cell in col:
            cell.border = full_border_thin

    for row in sheet.iter_rows(min_row=2, max_col=1):
        for cell in row:
            cell.font = font_not_bold
            cell.border = column_a_custom_border

    # Make adjustments to the end-of-table row
    cell1 = sheet["A" + end_of_table_row_str]
    cell1.border = top_border_thin

    cell2 = sheet["B" + end_of_table_row_str]
    cell2.font = font_bold
    cell2.border = full_border_medium

    cell3 = sheet["O" + end_of_table_row_str]
    cell3.font = font_bold
    cell3.border = full_border_medium

    # Format numbers
    number_format = "#,##0.00"
    for col in sheet.iter_cols(min_col=3, max_col=15):
        for cell in col:
            if isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.number_format = number_format

    workbook.save(path)


if __name__ == "__main__":
    style_spreadsheet("path", "2020 CRAW")
